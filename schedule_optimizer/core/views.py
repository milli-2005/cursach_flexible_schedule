# core/views.py
import secrets
import string
import json
import re
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import login, logout
from django.contrib import messages
from django.contrib.auth.forms import AuthenticationForm
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from django.conf import settings
from django.http import JsonResponse
from .error_utils import humanize_exception
from .services.rule_ai_parser import try_parse_rule_with_ai
from .email_utils import send_mail_with_fallback
from .models import *
from .forms import UserInvitationForm, UserProfileEditForm
from django.contrib.auth.models import User
import logging
from django.utils import timezone #для времени сброса пароля
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404
from .models import Schedule, UserProfile, WorkoutType

from django.shortcuts import render
from .models import Schedule, ShiftAssignment
from collections import defaultdict

from django.shortcuts import redirect
from django.contrib import messages



logger = logging.getLogger(__name__)

def is_manager(user):
    """
    Проверяет, является ли пользователь руководителем.
    """
    if not hasattr(user, 'profile'):
        return False
    return user.profile.role == 'manager'


def generate_random_password(length=12):
    """Генерирует случайный безопасный пароль."""
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    password = ''.join(secrets.choice(alphabet) for _ in range(length))
    return password

def send_user_invitation(user, raw_password):
    """Отправляет приглашение новому пользователю с паролем."""
    subject = 'Приглашение в систему планирования смен'
    # HTML сообщение
    html_message = render_to_string('core/emails/user_invitation.html', {
        'user': user,
        'raw_password': raw_password,
        'site_url': 'http://localhost:8000',  # Замените на ваш домен
        'login_url': 'http://localhost:8000/login/',
    })

    # !!! ВАЖНО: замените на ваш публичный URL (например, от ngrok).
    # site_url = getattr(settings, 'PUBLIC_SITE_URL', 'http://localhost:8000')  # Используем переменную из settings
    # html_message = render_to_string('core/emails/user_invitation.html', {
    #     'user': user,
    #     'raw_password': raw_password,
    #     'site_url': site_url,
    #     'login_url': f'{site_url}/login/',
    #     'change_password_url': f'{site_url}/profile/change-password/'  # Ссылка на смену пароля
    # })


    # Текстовое сообщение (для клиентов без поддержки HTML)
    plain_message = strip_tags(html_message)

    # Отправляем email
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [user.email], # Отправляем на email нового пользователя
        html_message=html_message,
        fail_silently=False,
    )



def is_admin(user):
    """Проверяет, является ли пользователь администратором."""
    if not hasattr(user, 'profile'):
        return False
    return user.profile.role == 'manager' or user.is_superuser


@login_required
@user_passes_test(is_admin)
def invite_user(request):
    """
    Страница для приглашения нового пользователя.
    Доступна только администраторам.
    """
    if request.method == 'POST':
        form = UserInvitationForm(request.POST)
        if form.is_valid():
            try:
                # Генерируем случайный пароль
                raw_password = generate_random_password()
                # Создаем пользователя
                user = User.objects.create_user(
                    username=form.cleaned_data['username'],
                    email=form.cleaned_data['email'],
                    password=raw_password, # Устанавливаем сгенерированный пароль
                    first_name=form.cleaned_data.get('first_name', ''),
                    last_name=form.cleaned_data.get('last_name', ''),
                )
                # Профиль создаётся автоматически сигналом в apps.py,
                # но мы должны обновить его атрибуты после создания
                profile = user.profile # Получаем связанный профиль
                profile.role = form.cleaned_data['role']
                # profile.position = form.cleaned_data.get('position', '')
                profile.phone = form.cleaned_data.get('phone', '')

                # Set temporary timestamp
                profile.invitation_timestamp = timezone.now()
                profile.save()

                # Отправляем приглашение на email
                try:
                    send_user_invitation(user, raw_password)
                    messages.success(
                        request,
                        f'Пользователь {user.username} успешно создан. '
                        f'Приглашение отправлено на {user.email}.'
                    )
                    logger.info(f'Администратор {request.user.username} создал пользователя {user.username}')
                except Exception as e:
                    # Если email не отправился, всё равно создаём пользователя
                    # и показываем пароль администратору
                    messages.warning(
                        request,
                        f'Пользователь {user.username} создан, но email не отправлен. '
                        f'Ошибка: {humanize_exception(e)}. Пароль пользователя: {raw_password}'
                    )
                    logger.error(f'Ошибка отправки email для пользователя {user.username}: {humanize_exception(e)}')

                return redirect('user_management') # Перенаправляем после успешного создания
            except Exception as e:
                messages.error(request, f'Ошибка при создании пользователя: {humanize_exception(e)}')
                logger.error(f'Ошибка создания пользователя: {humanize_exception(e)}')
    else:
        form = UserInvitationForm()

    return render(request, 'core/invite_user.html', {'form': form})



""" USER CRUD """
@login_required
@user_passes_test(is_admin)
def user_management(request):
    context = {}
    return render(request, 'core/dashboard/user_management.html', context)


@login_required
@user_passes_test(is_admin)
def reset_user_password(request, user_id):
    """Сброс пароля пользователя и отправка нового на email."""
    try:
        user = get_object_or_404(User, id=user_id)
        if request.method == 'POST':
            # Генерируем новый случайный пароль
            raw_password = generate_random_password()
            # Устанавливаем новый пароль
            user.set_password(raw_password)
            user.save()

            # Set temporary timestamp for reset
            profile = user.profile
            profile.invitation_timestamp = timezone.now()
            profile.save()

            # Отправляем email с новым паролем
            try:
                send_user_invitation(user, raw_password)
                messages.success(
                    request,
                    f'Новый пароль для пользователя {user.username} отправлен на {user.email}.'
                )
                logger.info(f'Администратор {request.user.username} сбросил пароль для {user.username}')
            except Exception as e:
                # Если email не отправился, показываем пароль администратору
                messages.warning(
                    request,
                    f'Пароль сброшен, но email не отправлен. '
                    f'Ошибка: {humanize_exception(e)}. Новый пароль: {raw_password}'
                )
            return redirect('user_management')

        context = {
            'user': user,
        }
        return render(request, 'profile/reset_password_confirm.html', context)
    except User.DoesNotExist:
        messages.error(request, 'Пользователь не найден.')
        return redirect('user_management')


def index(request):
    """Главная страница сайта."""
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('about')


def about(request):
    """Страница "О системе"."""
    return render(request, 'core/about.html')


@login_required
def chat_page(request):
    """Встроенный чат между пользователями."""
    return render(request, 'core/chat/chat.html')


def custom_login(request):
    """Кастомная страница входа в систему."""
    if request.user.is_authenticated:
        if hasattr(request.user, 'profile') and request.user.profile.invitation_timestamp:
            return redirect('change_password')
        return redirect('dashboard')

    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()

            # Проверяем, не истёк ли срок действия временного пароля
            if hasattr(user, 'profile') and user.profile.is_temporary_password_expired():
                messages.error(request,
                               "Временный пароль устарел. Пожалуйста, свяжитесь с администратором для получения нового.")
                # Важно: не логиним пользователя, если пароль истёк
                return render(request, 'core/login.html', {'form': form})

            login(request, user)
            if hasattr(user, 'profile') and user.profile.invitation_timestamp:
                messages.info(request, "Для продолжения работы сначала смените временный пароль.")
                return redirect('change_password')

            messages.success(request, f"Добро пожаловать, {user.username}!")
            return redirect('dashboard') # afo Перенаправляем на дашборд после входа
        else:
            messages.error(request, "Неверное имя пользователя или пароль.")
    else:
        form = AuthenticationForm()

    return render(request, 'core/login.html', {'form': form})


@login_required
def custom_logout(request):
    """Выход из системы."""
    logout(request)
    return redirect('index')


@login_required
def dashboard(request):
    """Личный кабинет пользователя."""
    user = request.user
    profile = user.profile
    context = {
        'user': user,
        'profile': profile,
    }

    # В зависимости от роли показываем разную информацию и перенаправляем или рендерим
    if profile.role == 'employee':
        # Для сотрудника показываем его смены и заявки
        try:
            employee_model = Employee.objects.get(user_profile=profile)
            context['employee'] = employee_model
            # Здесь можно добавить логику для получения смен сотрудника
        except Employee.DoesNotExist:
            pass
        return render(request, 'core/dashboard/dashboard_employee.html', context)
    elif profile.role == 'studio_admin':
        # Для администратора студии
        # Здесь можно добавить логику для получения графиков, отчетов и т.д.
        schedules = Schedule.objects.all()[:5]  # Пример
        context['schedules'] = schedules
        return render(request, 'core/dashboard/dashboard_studio_admin.html', context)
    elif profile.role == 'manager':
        # Для менеджера
        schedules = Schedule.objects.all()[:5]  # Пример
        context['schedules'] = schedules
        return render(request, 'core/dashboard/dashboard_manager.html', context)
    else:
        # На всякий случай, если роль неизвестна
        messages.error(request, "Неизвестная роль пользователя.")
        return redirect('index')




@login_required
def profile_view(request):
    user = request.user
    profile = user.profile
    employee, created = Employee.objects.get_or_create(user_profile=profile)

    context = {
        'user': user,
        'profile': profile,
        'employee': employee,
    }
    return render(request, 'core/profile/profile.html', context)



@login_required
def profile_edit(request):
    user = request.user
    profile = user.profile

    if request.method == 'POST':
        form = UserProfileEditForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Профиль успешно обновлён.")
            return redirect('profile_view')
    else:
        form = UserProfileEditForm(instance=profile)

    context = {'form': form}
    return render(request, 'core/profile/edit.html', context)



def change_password(request):
    """
    Смена пароля после регистрации по приглашению.
    Предполагается, что пользователь уже вошёл в систему с временным паролем.
    """
    from django.contrib.auth.forms import SetPasswordForm
    from django.contrib.auth import update_session_auth_hash  # Для обновления сессии

    user = request.user
    if not user.is_authenticated:
        messages.error(request, "Пожалуйста, войдите в систему, используя временный пароль из письма.")
        return redirect('login')

    # Проверяем, не истёк ли срок действия временного пароля при доступе к странице смены пароля
    if hasattr(user, 'profile') and user.profile.is_temporary_password_expired():
        messages.error(request, "Срок действия временного пароля истёк. Пожалуйста, свяжитесь с администратором для получения нового.")
        return redirect('login')  # Или на главную, если не хочет логиниться снова

    if request.method == 'POST':
        form = SetPasswordForm(user, request.POST)  # Передаём текущего пользователя
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # Важно: обновляем сессию, чтобы пользователь не вышел

            # Clear temporary timestamp after password change
            if hasattr(user, 'profile'):
                user.profile.invitation_timestamp = None
                user.profile.save()

            messages.success(request, "Ваш пароль был успешно изменён.")
            return redirect('profile_view')  # Перенаправляем на профиль после смены пароля
    else:
        form = SetPasswordForm(user)  # Передаём текущего пользователя

    context = {
        'form': form,
    }
    return render(request, 'core/profile/change_password.html', context)


@login_required
def optimization_view(request):
    """Страница оптимизации графиков."""

    # Проверяем права доступа
    if not hasattr(request.user, 'profile'):
        messages.error(request, "Профиль пользователя не найден.")
        return redirect('dashboard')

    user_profile = request.user.profile
    current_role = user_profile.current_role or user_profile.role

    # Только планировщики и админы могут использовать оптимизацию
    if current_role not in ['manager']:
        messages.error(request, "У вас нет доступа к этому разделу.")
        return redirect('dashboard')

    rules = OptimizationRule.objects.filter(is_active=True)
    context = {
        'rules': rules,
    }
    return render(request, 'core/optimization.html', context)



@login_required
def timeoff_requests(request):
    if not hasattr(request.user, 'profile'):
        messages.error(request, "Профиль пользователя не найден.")
        return redirect('dashboard')

    user_profile = request.user.profile
    current_role = user_profile.role # Берём роль из профиля

    # Права доступа могут отличаться
    # Сотрудник видит свои заявки, менеджер - все на согласование
    context = {
        'current_role': current_role,
    }
    return render(request, 'core/timeoff_requests.html', context)



@login_required
def shift_swaps(request):
    if request.user.profile.role != 'employee':
        return redirect('dashboard')

    shift_ids_str = request.GET.get('shift_ids')
    selected_shifts = []

    if shift_ids_str:
        try:
            shift_ids = [int(x) for x in shift_ids_str.split(',')]
            selected_shifts = ShiftAssignment.objects.filter(
                id__in=shift_ids,
                employee=request.user.profile
            )
        except ValueError:
            pass

    # Всегда загружаем список своих заявок
    my_requests = ShiftSwapRequest.objects.filter(
        from_employee__user_profile=request.user.profile
    ).select_related(
        'to_employee__user_profile__user'
    ).prefetch_related('shifts__shift_assignment').order_by('-created_at')

    context = {
        'selected_shifts': selected_shifts,
        'my_requests': my_requests,
    }

    if selected_shifts:
        context['shift_ids_json'] = json.dumps([s.id for s in selected_shifts])

    return render(request, 'core/swaps/shift_swaps.html', context)


@login_required
@user_passes_test(lambda u: u.profile.role in ['manager', 'studio_admin'])
def manager_swap_requests(request):
    """Страница для менеджера: просмотр и одобрение всех заявок"""
    swap_requests = ShiftSwapRequest.objects.select_related(
    'from_employee__user_profile__user',
    'to_employee__user_profile__user'
).prefetch_related(
    'shifts__shift_assignment'
).order_by('-created_at')

    context = {
        'swap_requests': swap_requests,
    }
    return render(request, 'core/swaps/manager_swap_requests.html', context)



def dashboard_employee(request):
    # Логика для дашборда сотрудника
    return render(request, 'core/dashboard_employee.html')

def dashboard_studio_admin(request):
    # Логика для дашборда админа студии
    return render(request, 'core/dashboard_studio_admin.html')

def dashboard_manager(request):
    # Логика для дашборда менеджера
    return render(request, 'core/dashboard_manager.html')

@login_required
def workout_types(request):
    """
    Страница управления типами занятий.
    Доступна только руководителю.
    """
    return render(request, 'core/workouts/workout_types.html')


""" === SCHEDULE === """
STUDIO_DAY_START_MIN = 9 * 60
STUDIO_DAY_END_MIN = 21 * 60
SLOT_WORK_MIN = 50
SLOT_BREAK_MIN = 10
STUDIO_LUNCH_START_MIN = 14 * 60
STUDIO_LUNCH_END_MIN = 16 * 60


def _generate_studio_slots():
    """Returns studio slots excluding lunch break 14:00-16:00."""
    slots = []
    current_time = STUDIO_DAY_START_MIN
    while current_time + SLOT_WORK_MIN <= STUDIO_DAY_END_MIN:
        start_min = current_time
        end_min = current_time + SLOT_WORK_MIN
        intersects_lunch = start_min < STUDIO_LUNCH_END_MIN and end_min > STUDIO_LUNCH_START_MIN
        if not intersects_lunch:
            start_str = f"{start_min // 60:02d}:{start_min % 60:02d}"
            end_str = f"{end_min // 60:02d}:{end_min % 60:02d}"
            slots.append((start_str, end_str))
        current_time = end_min + SLOT_BREAK_MIN
    return slots


DAY_NAME_TO_INDEX = {
    'понедельник': 0, 'пн': 0,
    'вторник': 1, 'вт': 1,
    'среда': 2, 'ср': 2,
    'четверг': 3, 'чт': 3,
    'пятница': 4, 'пт': 4,
    'суббота': 5, 'сб': 5,
    'воскресенье': 6, 'вс': 6,
}


def _normalize_rule_text(text: str) -> str:
    return re.sub(r'\s+', ' ', (text or '').strip().lower())


WORKOUT_CATEGORY_ALIASES = {
    'спокойн': 'calm',
    'calm': 'calm',
    'йога': 'calm',
    'стретч': 'calm',
    'растяж': 'calm',
    'пилатес': 'calm',
    'кардио': 'cardio',
    'cardio': 'cardio',
    'табата': 'cardio',
    'hiit': 'cardio',
    'силов': 'strength',
    'strength': 'strength',
    'power': 'strength',
    'танц': 'dance',
    'dance': 'dance',
    'бачата': 'dance',
    'стрип': 'dance',
    'восточ': 'dance',
}


def _extract_category_from_text(src: str):
    for key, value in WORKOUT_CATEGORY_ALIASES.items():
        if key in src:
            return value
    return None


def _parse_distribution_rule_text(text: str):
    src = _normalize_rule_text(text)
    if not src:
        return None, 'Введите текст правила.'

    # 1) "Табата не более 1 раза утром и 1 раза вечером в неделю"
    weekly_pattern = re.search(
        r'(?P<target>[а-яa-z0-9 \-_]+?)\s+.*?не более\s+(?P<morning>\d+)\s+раз.*?утр.*?(?P<evening>\d+)\s+раз.*?вечер.*?недел',
        src
    )
    if weekly_pattern:
        raw_target = weekly_pattern.group('target').strip(' "«»')
        target_category = _extract_category_from_text(raw_target)
        morning_max = int(weekly_pattern.group('morning'))
        evening_max = int(weekly_pattern.group('evening'))
        params = {
            'period': 'week',
            'buckets': [
                {'name': 'morning', 'start': '09:00', 'end': '14:00', 'max': morning_max},
                {'name': 'evening', 'start': '16:00', 'end': '21:00', 'max': evening_max},
            ],
        }
        if target_category:
            params.update({'target_mode': 'category', 'category': target_category})
            title = f'Лимит категории "{target_category}" по неделе'
        else:
            params.update({'target_mode': 'workout', 'workout_name': raw_target})
            title = f'Лимит "{raw_target}" по неделе'
        payload = {
            'rule_type': 'weekly_limit',
            'severity': 'hard',
            'name': title,
            'params_json': params
        }
        return payload, None

    # 1.1) "табата только 2 раза в неделю" / "не более 2 раз в неделю"
    total_week_pattern = re.search(
        r'(?P<target>[а-яa-z0-9 \-_]+?)\s+.*?(?:только|не более)\s+(?P<count>\d+)\s+раз\w*\s+.*?недел',
        src
    )
    if total_week_pattern:
        raw_target = total_week_pattern.group('target').strip(' "«»')
        target_category = _extract_category_from_text(raw_target)
        total_max = int(total_week_pattern.group('count'))
        params = {
            'period': 'week',
            'max_total': total_max,
        }
        if target_category:
            params.update({'target_mode': 'category', 'category': target_category})
            title = f'Лимит категории "{target_category}" за неделю'
        else:
            params.update({'target_mode': 'workout', 'workout_name': raw_target})
            title = f'Лимит "{raw_target}" за неделю'
        payload = {
            'rule_type': 'weekly_limit',
            'severity': 'hard',
            'name': title,
            'params_json': params
        }
        return payload, None

    # 1.2) "две одинаковые тренировки в один день утром/вечером нельзя"
    duplicate_day_pattern = (
        ('одинаков' in src or 'дубликат' in src) and
        ('один день' in src or 'в один день' in src or 'за день' in src) and
        ('нельзя' in src or 'запрет' in src or 'не став' in src)
    )
    if duplicate_day_pattern:
        payload = {
            'rule_type': 'daily_duplicate_limit',
            'severity': 'hard',
            'name': 'Запрет одинаковых тренировок в день (утро/вечер)',
            'params_json': {
                'scope': 'bucket',
                'max_per_bucket_per_day': 1,
                'buckets': [
                    {'name': 'morning', 'start': '09:00', 'end': '14:00'},
                    {'name': 'evening', 'start': '16:00', 'end': '21:00'},
                ],
            }
        }
        return payload, None

    # 2) "по понедельникам и средам допускаются две спокойные тренировки подряд"
    if 'спокойн' in src and 'подряд' in src and ('понедель' in src or 'сред' in src):
        weekdays = []
        for key, value in DAY_NAME_TO_INDEX.items():
            if key in src and value not in weekdays:
                weekdays.append(value)
        if not weekdays:
            weekdays = [0, 2]
        payload = {
            'rule_type': 'calm_consecutive',
            'severity': 'hard',
            'name': 'Спокойные подряд в выбранные дни',
            'params_json': {
                'weekdays': sorted(set(weekdays)),
                'max_consecutive': 2,
                'category': 'calm',
            }
        }
        return payload, None

    # 2.1) "не нужно ставить несколько силовых тренировок подряд"
    if 'силов' in src and 'подряд' in src:
        payload = {
            'rule_type': 'calm_consecutive',
            'severity': 'hard',
            'name': 'Запрет нескольких силовых подряд',
            'params_json': {
                'weekdays': [0, 1, 2, 3, 4, 5, 6],
                'max_consecutive': 1,
                'category': 'strength',
            }
        }
        return payload, None

    # 3) "силовые и кардио должны чередоваться"
    if 'силов' in src and 'кардио' in src and ('черед' in src):
        other_days = [1, 3, 4, 5, 6]
        payload = {
            'rule_type': 'alternation',
            'severity': 'hard',
            'name': 'Чередование силовых и кардио',
            'params_json': {
                'weekdays': other_days,
                'categories': ['strength', 'cardio'],
                'mode': 'strict_alternate',
            }
        }
        return payload, None

    return None, 'Не удалось распознать правило. Сейчас поддерживаются 4 шаблона из примеров.'


def _serialize_active_distribution_rules():
    rules = DistributionRule.objects.filter(is_active=True).order_by('priority', 'id')
    serialized = []
    for rule in rules:
        serialized.append({
            'id': rule.id,
            'name': rule.name,
            'rule_type': rule.rule_type,
            'severity': rule.severity,
            'params': rule.params_json or {},
        })
    return serialized


def _bucket_signature(bucket: dict) -> str:
    return f"{(bucket or {}).get('name','')}|{(bucket or {}).get('start','')}|{(bucket or {}).get('end','')}"


def _extract_weekly_limit_map(rule: DistributionRule):
    if rule.rule_type != 'weekly_limit':
        return {}
    params = rule.params_json or {}
    target_mode = params.get('target_mode') or ('category' if params.get('category') else 'workout')
    target_key = (params.get('category') or '').strip().lower() if target_mode == 'category' else _normalize_workout_name_for_rule(params.get('workout_name') or '')
    if not target_key:
        return {}
    result = {}
    for bucket in (params.get('buckets') or []):
        try:
            max_value = int(bucket.get('max', 0))
        except Exception:
            max_value = 0
        result[_bucket_signature(bucket)] = {
            'max': max_value,
            'bucket': bucket,
            'target_mode': target_mode,
            'target_key': target_key,
        }
    return result


def _extract_alternation_signature(rule: DistributionRule):
    if rule.rule_type != 'alternation':
        return None
    params = rule.params_json or {}
    categories = [str(x).strip().lower() for x in (params.get('categories') or []) if str(x).strip()]
    weekdays = sorted(set(int(x) for x in (params.get('weekdays') or []) if str(x).strip().isdigit()))
    if len(categories) < 2 or not weekdays:
        return None
    return {
        'categories': tuple(sorted(set(categories))),
        'weekdays': tuple(weekdays),
    }


def _build_distribution_rules_conflicts(rules):
    """
    Ищет явные и потенциальные противоречия между активными правилами.
    Возвращает список словарей для отображения на странице.
    """
    active_rules = [r for r in rules if r.is_active]
    conflicts = []

    for i in range(len(active_rules)):
        for j in range(i + 1, len(active_rules)):
            a = active_rules[i]
            b = active_rules[j]

            # 1) Явный конфликт: два weekly_limit на один и тот же таргет и одно окно,
            # но с разным max.
            a_week = _extract_weekly_limit_map(a)
            b_week = _extract_weekly_limit_map(b)
            if a_week and b_week:
                for bucket_key, a_data in a_week.items():
                    b_data = b_week.get(bucket_key)
                    if not b_data:
                        continue
                    if a_data['target_mode'] != b_data['target_mode'] or a_data['target_key'] != b_data['target_key']:
                        continue
                    if a_data['max'] != b_data['max']:
                        conflicts.append({
                            'level': 'hard',
                            'title': 'Противоречивые лимиты',
                            'rule_a': a,
                            'rule_b': b,
                            'description': (
                                f'Для одного и того же ограничения заданы разные лимиты '
                                f'в окне "{a_data["bucket"].get("name", "slot")}".'
                            ),
                            'how_to_fix': 'Оставьте один лимит или сделайте одинаковые значения max в обоих правилах.',
                        })

            # 2) Потенциальный конфликт: две alternation со схожими днями и общей категорией,
            # но разными наборами категорий.
            a_alt = _extract_alternation_signature(a)
            b_alt = _extract_alternation_signature(b)
            if a_alt and b_alt:
                weekdays_intersection = set(a_alt['weekdays']) & set(b_alt['weekdays'])
                categories_intersection = set(a_alt['categories']) & set(b_alt['categories'])
                a_cnt = len(set(a_alt['categories']))
                b_cnt = len(set(b_alt['categories']))
                mixed_scope = (a_cnt >= 3 and b_cnt == 2) or (b_cnt >= 3 and a_cnt == 2)

                if (
                    weekdays_intersection
                    and categories_intersection
                    and set(a_alt['categories']) != set(b_alt['categories'])
                    and not mixed_scope
                ):
                    conflicts.append({
                        'level': 'soft',
                        'title': 'Возможный конфликт чередования',
                        'rule_a': a,
                        'rule_b': b,
                        'description': (
                            'Для пересекающихся дней заданы разные пары категорий чередования. '
                            'Алгоритм может заполнять такие дни нестабильно.'
                        ),
                        'how_to_fix': 'Разведите правила по разным дням недели или оставьте одну пару категорий на один набор дней.',
                    })

                # 2.1) Усиленный конфликт: одно alternation общее (3+ категорий),
                # а второе более узкое (2 категории) на пересекающиеся дни.
                if weekdays_intersection and ((a_cnt >= 3 and b_cnt == 2) or (b_cnt >= 3 and a_cnt == 2)):
                    conflicts.append({
                        'level': 'hard',
                        'title': 'Противоречивые схемы чередования',
                        'rule_a': a,
                        'rule_b': b,
                        'description': (
                            'Широкое правило чередования (с 3+ категориями) пересекается с узким '
                            'правилом (2 категории) по тем же дням.'
                        ),
                        'how_to_fix': 'Оставьте одно правило чередования на эти дни или разделите дни между правилами.',
                    })

            # 3) Потенциальный конфликт приоритетов: два hard-правила одного типа и одного таргета.
            if a.rule_type == b.rule_type and a.severity == 'hard' and b.severity == 'hard':
                if a.rule_type == 'weekly_limit' and _extract_weekly_limit_map(a) and _extract_weekly_limit_map(b):
                    conflicts.append({
                        'level': 'soft',
                        'title': 'Перекрывающиеся жесткие weekly_limit',
                        'rule_a': a,
                        'rule_b': b,
                        'description': 'Два жестких weekly_limit могут дублировать друг друга и усложнять отладку.',
                        'how_to_fix': 'Объедините их в одно правило или понизьте жесткость/измените приоритет одного из них.',
                    })

    return conflicts


@login_required
@user_passes_test(is_manager)
def distribution_rules_page(request):
    rules = DistributionRule.objects.all().select_related('created_by').order_by('priority', 'id')
    conflicts = _build_distribution_rules_conflicts(list(rules))
    conflict_rule_ids = set()
    conflict_rule_hard_ids = set()
    conflict_rule_soft_ids = set()
    for c in conflicts:
        level = c.get('level') or 'soft'
        if c.get('rule_a'):
            rid = c['rule_a'].id
            conflict_rule_ids.add(rid)
            if level == 'hard':
                conflict_rule_hard_ids.add(rid)
            else:
                conflict_rule_soft_ids.add(rid)
        if c.get('rule_b'):
            rid = c['rule_b'].id
            conflict_rule_ids.add(rid)
            if level == 'hard':
                conflict_rule_hard_ids.add(rid)
            else:
                conflict_rule_soft_ids.add(rid)
    conflict_rule_soft_ids = conflict_rule_soft_ids - conflict_rule_hard_ids
    return render(
        request,
        'core/schedules/distribution_rules.html',
        {
            'rules': rules,
            'rules_conflicts': conflicts,
            'rules_conflicts_count': len(conflicts),
            'conflict_rule_ids': sorted(conflict_rule_ids),
            'conflict_rule_hard_ids': sorted(conflict_rule_hard_ids),
            'conflict_rule_soft_ids': sorted(conflict_rule_soft_ids),
        }
    )


@login_required
@user_passes_test(is_manager)
def api_parse_distribution_rule(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = {}
    text = (payload.get('text') or '').strip()
    if not text:
        return JsonResponse({'success': False, 'error': 'Введите текст правила.'}, status=400)

    ai_result = try_parse_rule_with_ai(text)
    if ai_result.get('success'):
        return JsonResponse({
            'success': True,
            'parsed': ai_result['parsed'],
            'source': 'ai',
            'explanation': ai_result.get('explanation') or 'Распознано с помощью ИИ.',
            'confidence': ai_result.get('confidence', 0.85),
        })

    parsed, error = _parse_distribution_rule_text(text)
    if error:
        ai_error = ai_result.get('error')
        suffix = f" AI: {ai_error}" if ai_error else ""
        return JsonResponse({'success': False, 'error': f'{error}{suffix}'}, status=400)
    return JsonResponse({
        'success': True,
        'parsed': parsed,
        'source': 'fallback_regex',
        'explanation': 'Распознано резервным шаблонным парсером.',
        'confidence': 0.72,
    })


@login_required
@user_passes_test(is_manager)
def api_save_distribution_rule(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = {}
    source_text = (payload.get('source_text') or '').strip()
    parsed = payload.get('parsed') or {}
    if not source_text:
        return JsonResponse({'success': False, 'error': 'Пустой текст правила.'}, status=400)
    if not parsed or not parsed.get('rule_type'):
        return JsonResponse({'success': False, 'error': 'Нет распознанных данных правила.'}, status=400)

    rule = DistributionRule.objects.create(
        name=(payload.get('name') or parsed.get('name') or source_text[:180]).strip()[:200],
        source_text=source_text,
        rule_type=parsed.get('rule_type'),
        severity=parsed.get('severity') if parsed.get('severity') in {'hard', 'soft'} else 'hard',
        params_json=parsed.get('params_json') or {},
        is_active=bool(payload.get('is_active', True)),
        priority=int(payload.get('priority', 100) or 100),
        created_by=request.user,
    )
    conflicts = _build_distribution_rules_conflicts(
        list(DistributionRule.objects.all().order_by('priority', 'id'))
    )
    return JsonResponse({
        'success': True,
        'rule_id': rule.id,
        'conflicts_count': len(conflicts),
    })


@login_required
@user_passes_test(is_manager)
def api_toggle_distribution_rule(request, rule_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    rule = get_object_or_404(DistributionRule, id=rule_id)
    rule.is_active = not rule.is_active
    rule.save(update_fields=['is_active', 'updated_at'])
    return JsonResponse({'success': True, 'is_active': rule.is_active})


@login_required
@user_passes_test(is_manager)
def api_delete_distribution_rule(request, rule_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    rule = get_object_or_404(DistributionRule, id=rule_id)
    rule.delete()
    return JsonResponse({'success': True})


@login_required
@user_passes_test(is_manager)
def api_update_distribution_rule(request, rule_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    rule = get_object_or_404(DistributionRule, id=rule_id)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = {}

    name = (payload.get('name') or '').strip()
    severity = (payload.get('severity') or '').strip()
    priority_raw = payload.get('priority', rule.priority)

    if name:
        rule.name = name[:200]
    if severity in {'hard', 'soft'}:
        rule.severity = severity
    try:
        rule.priority = max(1, int(priority_raw))
    except Exception:
        pass

    rule.save(update_fields=['name', 'severity', 'priority', 'updated_at'])
    return JsonResponse({
        'success': True,
        'rule': {
            'id': rule.id,
            'name': rule.name,
            'severity': rule.severity,
            'priority': rule.priority,
        }
    })


def _infer_category_from_name(workout_name: str) -> str:
    n = (workout_name or '').lower()
    if any(x in n for x in ['табата', 'кардио', 'cardio', 'hiit']):
        return 'cardio'
    if any(x in n for x in ['сил', 'strength', 'power']):
        return 'strength'
    if any(x in n for x in ['dance', 'танц', 'bachata', 'восточ', 'стрип']):
        return 'dance'
    if any(x in n for x in ['stretch', 'растяж', 'йог', 'calm', 'спокой']):
        return 'calm'
    return 'other'


def _normalize_workout_name_for_rule(name: str) -> str:
    n = (name or '').strip().lower()
    aliases = {
        'табата': 'tabata',
        'стретчинг': 'stretching',
        'растяжка': 'stretching',
        'бачата': 'bachata',
        'силовые': 'strength',
        'кардио': 'cardio',
    }
    return aliases.get(n, n)


def _time_in_bucket(start_time, bucket):
    st = start_time.strftime('%H:%M')
    return (bucket.get('start') or '00:00') <= st < (bucket.get('end') or '23:59')


@login_required
@user_passes_test(is_manager)
def api_test_distribution_rules(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Метод не поддерживается.'}, status=405)
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = {}

    start_raw = (payload.get('start_date') or '').strip()
    end_raw = (payload.get('end_date') or '').strip()
    if not start_raw or not end_raw:
        return JsonResponse({'success': False, 'error': 'Укажите период.'}, status=400)
    try:
        start_date = datetime.strptime(start_raw, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_raw, '%Y-%m-%d').date()
    except Exception:
        return JsonResponse({'success': False, 'error': 'Некорректный формат даты.'}, status=400)
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    rules = list(DistributionRule.objects.filter(is_active=True).order_by('priority', 'id'))
    assignments = list(
        ShiftAssignment.objects.filter(
            date__gte=start_date,
            date__lte=end_date,
            workout_type__isnull=False,
        ).select_related('workout_type', 'employee__user')
    )
    assignments.sort(key=lambda a: (a.date, a.start_time))

    violations = []
    weekly_counts = {}
    weekly_total_counts = {}
    daily_bucket_workout_counts = {}
    calm_streaks = {}
    prev_category = {}

    for a in assignments:
        workout_name = a.workout_type.name if a.workout_type_id else ''
        category = (a.workout_type.category if getattr(a.workout_type, 'category', None) else _infer_category_from_name(workout_name))
        weekday = a.date.weekday()
        day_key = a.date.isoformat()
        week_key = f"{a.date.isocalendar().year}-{a.date.isocalendar().week}"

        for rule in rules:
            params = rule.params_json or {}
            if rule.rule_type == 'weekly_limit':
                target_mode = params.get('target_mode') or ('category' if params.get('category') else 'workout')
                target_workout = _normalize_workout_name_for_rule(params.get('workout_name') or '')
                target_category = (params.get('category') or '').strip()
                workout_norm = _normalize_workout_name_for_rule(workout_name)
                is_match = False
                if target_mode == 'category':
                    is_match = bool(target_category and category == target_category)
                else:
                    is_match = bool(target_workout and target_workout in workout_norm)
                if is_match:
                    total_key = f"{rule.id}|{week_key}|total"
                    weekly_total_counts[total_key] = weekly_total_counts.get(total_key, 0) + 1
                    if params.get('max_total') is not None and weekly_total_counts[total_key] > int(params.get('max_total', 0)):
                        violations.append({
                            'rule': rule.name,
                            'date': a.date.strftime('%d.%m.%Y'),
                            'time': a.start_time.strftime('%H:%M'),
                            'workout': workout_name,
                            'employee': a.employee.user.username,
                            'reason': 'Превышен общий недельный лимит',
                        })
                    for b in (params.get('buckets') or []):
                        if _time_in_bucket(a.start_time, b):
                            key = f"{rule.id}|{week_key}|{b.get('name','bucket')}"
                            weekly_counts[key] = weekly_counts.get(key, 0) + 1
                            if weekly_counts[key] > int(b.get('max', 0)):
                                violations.append({
                                    'rule': rule.name,
                                    'date': a.date.strftime('%d.%m.%Y'),
                                    'time': a.start_time.strftime('%H:%M'),
                                    'workout': workout_name,
                                    'employee': a.employee.user.username,
                                    'reason': f'Превышен лимит "{b.get("name", "bucket")}" за неделю',
                                })
            elif rule.rule_type == 'calm_consecutive':
                weekdays = params.get('weekdays') or []
                max_consecutive = int(params.get('max_consecutive', 2))
                expected = params.get('category', 'calm')
                if weekday in weekdays:
                    if category == expected:
                        calm_streaks[day_key] = calm_streaks.get(day_key, 0) + 1
                        if calm_streaks[day_key] > max_consecutive:
                            violations.append({
                                'rule': rule.name,
                                'date': a.date.strftime('%d.%m.%Y'),
                                'time': a.start_time.strftime('%H:%M'),
                                'workout': workout_name,
                                'employee': a.employee.user.username,
                                'reason': 'Слишком много спокойных подряд',
                            })
                    else:
                        calm_streaks[day_key] = 0
            elif rule.rule_type == 'alternation':
                weekdays = params.get('weekdays') or []
                categories = params.get('categories') or ['strength', 'cardio']
                if weekday in weekdays and category in categories:
                    if prev_category.get(day_key) == category:
                        violations.append({
                            'rule': rule.name,
                            'date': a.date.strftime('%d.%m.%Y'),
                            'time': a.start_time.strftime('%H:%M'),
                            'workout': workout_name,
                            'employee': a.employee.user.username,
                            'reason': 'Нарушено чередование категорий',
                        })
                    prev_category[day_key] = category
            elif rule.rule_type == 'daily_duplicate_limit':
                buckets = params.get('buckets') or []
                max_per_bucket_per_day = int(params.get('max_per_bucket_per_day', 1))
                workout_norm = _normalize_workout_name_for_rule(workout_name)
                for b in buckets:
                    if _time_in_bucket(a.start_time, b):
                        key = f"{rule.id}|{day_key}|{b.get('name','bucket')}|{workout_norm}"
                        daily_bucket_workout_counts[key] = daily_bucket_workout_counts.get(key, 0) + 1
                        if daily_bucket_workout_counts[key] > max_per_bucket_per_day:
                            violations.append({
                                'rule': rule.name,
                                'date': a.date.strftime('%d.%m.%Y'),
                                'time': a.start_time.strftime('%H:%M'),
                                'workout': workout_name,
                                'employee': a.employee.user.username,
                                'reason': 'Одинаковая тренировка повторяется в одном окне дня',
                            })

    return JsonResponse({
        'success': True,
        'period': {
            'start': start_date.isoformat(),
            'end': end_date.isoformat(),
        },
        'rules_count': len(rules),
        'violations_count': len(violations),
        'violations': violations[:150],
    })


@login_required
def create_schedule_view(request):
    employee_models_with_workouts = Employee.objects.select_related('user_profile__user').prefetch_related('workout_types').filter(
        user_profile__role='employee',
        user_profile__user__is_active=True,
        workout_types__isnull=False,
    ).distinct()
    employees = UserProfile.objects.filter(
        id__in=employee_models_with_workouts.values_list('user_profile_id', flat=True)
    ).select_related('user')
    workout_types = WorkoutType.objects.all()

    # Генерация слотов (с учетом обеда 14:00-16:00)
    slots = _generate_studio_slots()

    # Build days for next week
    today = datetime.today()
    next_monday = today + timedelta(days=(7 - today.weekday()))
    current_days = [next_monday.date() + timedelta(days=i) for i in range(7)]

    # Строки для JS и шаблона
    date_strings = [d.strftime('%Y-%m-%d') for d in current_days]

    # Загрузка доступности, подготовка данных для быстрой проверки в JS
    availabilities = Availability.objects.filter(
        employee__in=employees,
        date__in=current_days,
        is_available=True,
    )

    # Создаём SET для быстрой проверки в JS: "emp_id,date,time"
    availability_set = set()
    for a in availabilities:
        key = f"{a.employee.id},{a.date.strftime('%Y-%m-%d')},{a.start_time.strftime('%H:%M')}"
        availability_set.add(key)
        
     # Получаем всех сотрудников с их направлениями
    employees_with_workouts = []
    for emp in employee_models_with_workouts:
        workouts = list(emp.workout_types.values('id', 'name', 'category'))
        employees_with_workouts.append({
            'id': emp.user_profile.id,
            'username': emp.user_profile.user.username,
            'workout_types': workouts
        })

    missing_workout_profiles = UserProfile.objects.filter(
        role='employee',
        user__is_active=True,
    ).exclude(
        id__in=employees.values_list('id', flat=True)
    ).select_related('user')

    context = {
        'employees': employees,
        'workout_types': workout_types,
        'slots': slots,
        'days': current_days,
        'date_strings': date_strings,
        'availability_set_json': json.dumps(list(availability_set)),
        'employees_with_workouts_json': json.dumps(employees_with_workouts),
        'workout_types_json': json.dumps([
            {'id': wt.id, 'name': wt.name, 'category': wt.category} for wt in WorkoutType.objects.all()
        ]),
        'missing_workout_employees': [
            p.user.get_full_name().strip() or p.user.username for p in missing_workout_profiles
        ],
        'distribution_rules_json': json.dumps(_serialize_active_distribution_rules(), ensure_ascii=False),
    }
    
    return render(request, 'core/schedules/create_schedule.html', context)




from django.core.paginator import Paginator

@login_required
def schedule_view(request):
    from django.db.models import Q

    if not hasattr(request.user, 'profile'):
        messages.error(request, "Профиль пользователя не найден.")
        return redirect('dashboard')

    # Параметры
    try:
        page_size = int(request.GET.get('page_size', 6))
    except (TypeError, ValueError):
        page_size = 6
    if page_size not in [6, 10, 20, 50]:
        page_size = 6

    sort_by = request.GET.get('sort', '-start_date')  # по умолчанию — новые сверху
    query = request.GET.get('q', '').strip()
    status_filter = request.GET.get('status', '').strip()
    creator_filter = request.GET.get('creator', '').strip()
    period_filter = request.GET.get('period', '').strip()
    approval_filter = request.GET.get('approval', '').strip()

    # Валидация поля сортировки
    valid_sort_fields = [
        'name', '-name',
        'start_date', '-start_date',
        'end_date', '-end_date',
        'status', '-status',
        'created_at', '-created_at',
    ]
    if sort_by not in valid_sort_fields:
        sort_by = '-start_date'

    # Запрос с сортировкой
    schedules = Schedule.objects.all().select_related('created_by').prefetch_related('approvals')

    if query:
        schedules = schedules.filter(
            Q(name__icontains=query) |
            Q(created_by__username__icontains=query) |
            Q(created_by__first_name__icontains=query) |
            Q(created_by__last_name__icontains=query)
        )

    if status_filter in {'draft', 'pending', 'approved'}:
        schedules = schedules.filter(status=status_filter)

    if creator_filter.isdigit():
        schedules = schedules.filter(created_by_id=int(creator_filter))

    today = timezone.localdate()
    if period_filter == 'current':
        schedules = schedules.filter(start_date__lte=today, end_date__gte=today)
    elif period_filter == 'upcoming':
        schedules = schedules.filter(start_date__gt=today)
    elif period_filter == 'past':
        schedules = schedules.filter(end_date__lt=today)

    schedules = schedules.order_by(sort_by)

    total_employees = UserProfile.objects.filter(role='employee').count()

    # Аннотации
    schedules_with_stats = []
    for s in schedules:
        approved_count = s.approvals.filter(approved=True).count()
        rejected_count = s.approvals.filter(approved=False).count()
        responded_count = s.approvals.filter(approved__isnull=False).count()

        schedules_with_stats.append({
            'schedule': s,
            'total_employees': total_employees,
            'approved_count': approved_count,
            'rejected_count': rejected_count,
            'responded_count': responded_count,
        })

    # Фильтр по состоянию согласования (после подсчёта статистики)
    if approval_filter in {'not_reviewed', 'partially_reviewed', 'fully_reviewed', 'with_rejections'}:
        filtered_stats = []
        for item in schedules_with_stats:
            responded = item['responded_count']
            rejected = item['rejected_count']

            if approval_filter == 'not_reviewed' and responded == 0:
                filtered_stats.append(item)
            elif approval_filter == 'partially_reviewed' and 0 < responded < total_employees:
                filtered_stats.append(item)
            elif approval_filter == 'fully_reviewed' and responded >= total_employees and total_employees > 0:
                filtered_stats.append(item)
            elif approval_filter == 'with_rejections' and rejected > 0:
                filtered_stats.append(item)
        schedules_with_stats = filtered_stats

    # Пагинация
    paginator = Paginator(schedules_with_stats, page_size)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    creators = User.objects.filter(
        id__in=Schedule.objects.exclude(created_by__isnull=True).values_list('created_by_id', flat=True).distinct()
    ).order_by('last_name', 'first_name', 'username')

    context = {
        'schedules_with_stats': page_obj,
        'page_obj': page_obj,
        'page_size': page_size,
        'current_sort': sort_by,
        'filter_q': query,
        'filter_status': status_filter,
        'filter_creator': creator_filter,
        'filter_period': period_filter,
        'filter_approval': approval_filter,
        'creators': creators,
    }
    return render(request, 'core/schedules/schedule_list.html', context)



@login_required
def schedule_detail(request, schedule_id):
    schedule = get_object_or_404(Schedule, id=schedule_id)
    schedule_versions = schedule.versions.select_related('created_by').order_by('-version_number')[:30]

    # === 1. Генерация дней из графика ===
    days = []
    current_date = schedule.start_date
    while current_date <= schedule.end_date:
        days.append(current_date)
        current_date += timedelta(days=1)

    # === 2. Генерация временных слотов (9:00–21:00, без обеда 14:00-16:00) ===
    all_slots = [f"{start}–{end}" for start, end in _generate_studio_slots()]

    # === 3. Load all assignments in one query ===
    assignments = ShiftAssignment.objects.filter(
        schedule=schedule,
        date__in=days
    ).select_related('employee__user', 'workout_type')

    if (
        hasattr(request.user, 'profile')
        and request.user.profile.role in ['manager', 'studio_admin']
        and not schedule_versions
    ):
        bootstrap_version = ScheduleVersion.objects.create(
            schedule=schedule,
            version_number=1,
            schedule_name=schedule.name,
            created_by=request.user,
            change_source='bootstrap',
            change_note='Базовая версия для существующего графика',
        )
        snapshots = [
            ScheduleVersionAssignment(
                schedule_version=bootstrap_version,
                employee_id=a.employee_id,
                workout_type_id=a.workout_type_id,
                date=a.date,
                start_time=a.start_time,
                end_time=a.end_time,
            )
            for a in assignments
            if a.employee_id
        ]
        if snapshots:
            ScheduleVersionAssignment.objects.bulk_create(snapshots)
        schedule_versions = schedule.versions.select_related('created_by').order_by('-version_number')[:30]

    # === 4. Создание словаря: {(дата, время_начала): assignment} ===
    assignment_dict = {}
    for a in assignments:
        time_key = a.start_time.strftime('%H:%M')
        key = (a.date, time_key)
        assignment_dict[key] = a

    # === 5. Построение таблицы ===
    table_data = []
    for slot in all_slots:
        row = {'time_slot': slot, 'cells': []}
        start_time_str = slot.split('–')[0]

        for day in days:
            key = (day, start_time_str)
            assignment = assignment_dict.get(key)
            row['cells'].append({'assignment': assignment, 'date': day})
        table_data.append(row)

    # === 6. Данные для утверждения (только для сотрудников) ===
    approval_for_user = None
    approval_slots_for_user = []
    if request.user.is_authenticated and hasattr(request.user, 'profile'):
        if request.user.profile.role == 'employee':
            try:
                approval_for_user = ScheduleApproval.objects.get(
                    schedule=schedule,
                    employee=request.user.profile
                )
            except ScheduleApproval.DoesNotExist:
                approval_for_user = None
            user_shifts = ShiftAssignment.objects.filter(
                schedule=schedule,
                employee=request.user.profile,
            ).select_related('workout_type').order_by('date', 'start_time')
            approval_slots_for_user = [
                {
                    'date': sh.date.strftime('%Y-%m-%d'),
                    'date_label': sh.date.strftime('%d.%m.%Y'),
                    'start_time': sh.start_time.strftime('%H:%M'),
                    'time_slot': f"{sh.start_time.strftime('%H:%M')}-{sh.end_time.strftime('%H:%M')}",
                    'workout_name': sh.workout_type.name if sh.workout_type else 'Занятие',
                }
                for sh in user_shifts
            ]

    employee_models_with_workouts = Employee.objects.select_related('user_profile__user').prefetch_related('workout_types').filter(
        user_profile__role='employee',
        user_profile__user__is_active=True,
        workout_types__isnull=False,
    ).distinct()
    employees_with_workouts = []
    for emp in employee_models_with_workouts:
        workouts = list(emp.workout_types.values('id', 'name', 'category'))
        employees_with_workouts.append({
            'id': emp.user_profile.id,
            'username': emp.user_profile.user.username,
            'workout_types': workouts,
        })

    availabilities = Availability.objects.filter(
        employee_id__in=employee_models_with_workouts.values_list('user_profile_id', flat=True),
        date__in=days,
        is_available=True,
    )
    availability_set = set()
    for a in availabilities:
        key = f"{a.employee.id},{a.date.strftime('%Y-%m-%d')},{a.start_time.strftime('%H:%M')}"
        availability_set.add(key)

    editable_employee_ids = list(
        employee_models_with_workouts.values_list('user_profile_id', flat=True)
    )

    schedule_edit_locked = False
    can_edit_schedule = (
        hasattr(request.user, 'profile')
        and request.user.profile.role in ['manager', 'studio_admin']
    )

    manager_rejections = []
    if hasattr(request.user, 'profile') and request.user.profile.role in ['manager', 'studio_admin']:
        rejected_approvals = (
            ScheduleApproval.objects
            .filter(schedule=schedule, approved=False)
            .select_related('employee__user')
            .order_by('-responded_at', '-id')
        )
        for appr in rejected_approvals:
            slots = appr.rejection_slots_json or []
            manager_rejections.append({
                'approval_id': appr.id,
                'employee_id': appr.employee_id,
                'employee_name': appr.employee.user.get_full_name() or appr.employee.user.username,
                'responded_at': appr.responded_at,
                'comment': appr.comment or '',
                'slots': slots,
            })

    context = {
        'schedule': schedule,
        'schedule_versions': schedule_versions,
        'days': days,
        'date_strings': [d.strftime('%Y-%m-%d') for d in days],
        'employees': UserProfile.objects.filter(
            id__in=editable_employee_ids
        ).select_related('user'),
        'editable_employee_ids': editable_employee_ids,
        'workout_types': WorkoutType.objects.all(),
        'employees_with_workouts_json': json.dumps(employees_with_workouts),
        'workout_types_json': json.dumps([
            {'id': wt.id, 'name': wt.name, 'category': wt.category} for wt in WorkoutType.objects.all()
        ]),
        'availability_set_json': json.dumps(list(availability_set)),
        'table_data': table_data,
        'approval_for_user': approval_for_user,
        'approval_slots_for_user_json': json.dumps(approval_slots_for_user, ensure_ascii=False),
        'manager_rejections': manager_rejections,
        'schedule_edit_locked': schedule_edit_locked,
        'can_edit_schedule': can_edit_schedule,
    }
    return render(request, 'core/schedules/view_schedule.html', context)




from collections import defaultdict

@login_required
def edit_schedule_view(request, schedule_id):
    schedule = get_object_or_404(Schedule, id=schedule_id)

    # Генерация слотов (9:00–21:00, без обеда 14:00-16:00)
    slots = [f"{start} – {end}" for start, end in _generate_studio_slots()]

    # Генерация дней из графика
    from datetime import timedelta
    days = []
    current_date = schedule.start_date
    while current_date <= schedule.end_date:
        days.append(current_date)
        current_date += timedelta(days=1)

    date_strings = [d.strftime('%Y-%m-%d') for d in days]

    employee_models_with_workouts = Employee.objects.select_related('user_profile__user').prefetch_related('workout_types').filter(
        user_profile__role='employee',
        user_profile__user__is_active=True,
        workout_types__isnull=False,
    ).distinct()
    employees = UserProfile.objects.filter(
        id__in=employee_models_with_workouts.values_list('user_profile_id', flat=True)
    ).select_related('user')
    workout_types = WorkoutType.objects.all()

    # Текущие назначения
    assignments = ShiftAssignment.objects.filter(schedule=schedule).select_related('employee', 'workout_type')
    assignment_dict = defaultdict(dict)
    for a in assignments:
        time_key = a.start_time.strftime('%H:%M')
        assignment_dict[a.date][time_key] = a

    # === Load local data for rendering ===
    availabilities = Availability.objects.filter(
        employee__in=employees,
        date__in=days,
        is_available=True,
    )
    availability_set = set()
    for a in availabilities:
        key = f"{a.employee.id},{a.date.strftime('%Y-%m-%d')},{a.start_time.strftime('%H:%M')}"
        availability_set.add(key)

    context = {
        'schedule': schedule,
        'slots': slots,
        'days': days,
        'date_strings': date_strings,  # ← для JS
         'date_strings_json': json.dumps([d.strftime('%Y-%m-%d') for d in days]),
        'employees': employees,
        'workout_types': workout_types,
        'assignment_dict': dict(assignment_dict),
        'availability_set_json': json.dumps(list(availability_set)),  # ← для JS
    }
    return render(request, 'core/schedules/edit_schedule.html', context)



@login_required
@user_passes_test(is_manager)
def delete_schedule_view(request, schedule_id):
    schedule = get_object_or_404(Schedule, id=schedule_id)
    if request.method == "POST":
        schedule_name = schedule.name
        schedule.delete()
        messages.success(request, f'График "{schedule_name}" успешно удалён.')
        return redirect('schedule_view')  # Перенаправление на список графиков
    # Если кто-то попытается GET — перенаправим на просмотр
    return redirect('view_schedule', schedule_id=schedule_id)




""" === EMPLOYEE SCHEDULE === """
from datetime import date, timedelta
import calendar
import json
from django.utils.html import escapejs
from django.utils import timezone

@login_required
def employee_schedule(request):
    if not hasattr(request.user, 'profile'):
        return redirect('dashboard')

    today = date.today()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
    except (TypeError, ValueError):
        year, month = today.year, today.month

    if month < 1:
        year -= 1
        month = 12
    elif month > 12:
        year += 1
        month = 1

    first_day = date(year, month, 1)
    if month == 12:
        last_day = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        last_day = date(year, month + 1, 1) - timedelta(days=1)

    all_dates = [first_day + timedelta(days=i) for i in range((last_day - first_day).days + 1)]

    assignments = ShiftAssignment.objects.filter(
        employee__user=request.user,
        date__gte=first_day,
        date__lte=last_day
    ).select_related('workout_type').order_by('date', 'start_time')

    shifts_by_date = {}
    for d in all_dates:
        shifts_by_date[d] = []
    for shift in assignments:
        shifts_by_date[shift.date].append(shift)

    # Подготовка JSON для JS
    shifts_json = {}
    for d, shifts in shifts_by_date.items():
        shifts_json[d.isoformat()] = [
            {
                'id': shift.id,
                'workout': escapejs(shift.workout_type.name if shift.workout_type else 'Работа'),
                'start': shift.start_time.strftime('%H:%M') if shift.start_time else '',
                'end': shift.end_time.strftime('%H:%M') if shift.end_time else '',
                'is_past': shift.date < today,
            }
            for shift in shifts
        ]

    cal = calendar.monthcalendar(year, month)
    weeks = []
    for week in cal:
        week_days = []
        for day in week:
            if day == 0:
                week_days.append(None)
            else:
                d = date(year, month, day)
                week_days.append({
                    'date': d,
                    'has_shift': len(shifts_by_date.get(d, [])) > 0,
                    'is_today': d == today,
                    'is_past': d < today,
                })
        weeks.append(week_days)

    # === Графики на утверждение ===
    pending_approvals = []
    if request.user.profile.role == 'employee':
        today_local = timezone.localdate()
        days_to_next_monday = (7 - today_local.weekday()) % 7
        if days_to_next_monday == 0:
            days_to_next_monday = 7
        next_week_start = today_local + timedelta(days=days_to_next_monday)
        next_week_end = next_week_start + timedelta(days=6)

        pending_approvals = ScheduleApproval.objects.filter(
            employee=request.user.profile,
            approved__isnull=True,
            schedule__status='pending',
            schedule__start_date__lte=next_week_end,
            schedule__end_date__gte=next_week_start,
        ).select_related('schedule').order_by('schedule__start_date', 'schedule__id')

    context = {
        'current_year': year,
        'current_month': month,
        'month_name': first_day.strftime('%B %Y'),
        'weeks': weeks,
        'shifts_by_date': shifts_by_date,
        'today': today,
        'shifts_json': json.dumps(shifts_json, ensure_ascii=False),
        'pending_approvals': pending_approvals,  # ← Теперь переменная определена!
    }

    return render(request, 'core/schedules/employee_schedule.html', context)




""" === ДОСТУПНОСТЬ === """

# core/views.py
from datetime import datetime, timedelta
@login_required
def my_availability(request):
    user_profile = request.user.profile
    if user_profile.role != 'employee':
        messages.error(request, "Доступно только для сотрудников.")
        return redirect('dashboard')

    # === POST handling: save data ===
    if request.method == "POST":
        week_str = request.POST.get('selected_week')
        if week_str:
            try:
                week_start = datetime.strptime(week_str, '%Y-%m-%d').date()
                if week_start.weekday() != 0:
                    week_start = week_start - timedelta(days=week_start.weekday())
            except (ValueError, TypeError):
                messages.error(request, "Неверный формат даты.")
                return redirect('my_availability')
        else:
            today = datetime.today()
            days_ahead = (7 - today.weekday()) % 7
            if days_ahead == 0:
                days_ahead = 7
            week_start = today + timedelta(days=days_ahead)

        # Генерация дней
        current_days = [week_start + timedelta(days=i) for i in range(7)]
        date_strings = [d.strftime('%Y-%m-%d') for d in current_days]

        # Слоты (с учетом обеда 14:00-16:00)
        slots = _generate_studio_slots()

        print("=== POST KEYS ===")
        print(list(request.POST.keys()))
        print("=== EXPECTED SAMPLE ===")
        print(f"Sample key: {date_strings[0]}_{slots[0][0]}")

        # Удаление старых записей
        Availability.objects.filter(
            employee=user_profile,
            date__in=current_days
        ).delete()

        # Сохранение новых
        new_records = []
        for day_str in date_strings:
            for slot_start, slot_end in slots:
                key = f"{day_str}_{slot_start}"
                if request.POST.get(key) == 'on':  # <- enabled
                    date_obj = datetime.strptime(day_str, '%Y-%m-%d').date()
                    start_time = datetime.strptime(slot_start, '%H:%M').time()
                    end_time = datetime.strptime(slot_end, '%H:%M').time()
                    new_records.append(Availability(
                        employee=user_profile,
                        date=date_obj,
                        start_time=start_time,
                        end_time=end_time,
                        is_available=True
                    ))

        if new_records:
            Availability.objects.bulk_create(new_records)
            messages.success(request, "Доступность успешно сохранена!")
        else:
            messages.info(request, "Доступность не указана.")

        return redirect(f"{request.path}?week={week_start.strftime('%Y-%m-%d')}")

    # === GET handling: render form ===
    today = datetime.today()
    days_ahead = (7 - today.weekday()) % 7
    if days_ahead == 0:
        days_ahead = 7
    default_week_start = today + timedelta(days=days_ahead)

    week_start = default_week_start
    week_param = request.GET.get('week')
    if week_param:
        try:
            parsed_date = datetime.strptime(week_param, '%Y-%m-%d').date()
            week_start = parsed_date - timedelta(days=parsed_date.weekday())
        except (ValueError, TypeError):
            pass

    # Генерация дней (без спискового включения — через цикл)
    current_days = []
    for i in range(7):
        current_days.append(week_start + timedelta(days=i))
    date_strings = [d.strftime('%Y-%m-%d') for d in current_days]

    # Слоты (с учетом обеда 14:00-16:00)
    slots = _generate_studio_slots()

    # Загрузка данных
    availabilities = Availability.objects.filter(
        employee=user_profile,
        date__in=current_days
    )
    checked_keys = set()
    for a in availabilities:
        key = f"{a.date.strftime('%Y-%m-%d')}_{a.start_time.strftime('%H:%M')}"
        checked_keys.add(key)

    last_updated = availabilities.order_by('-updated_at').first()

    # === Previous-week data for JS ===
    prev_week_start = week_start - timedelta(weeks=1)
    prev_avail = Availability.objects.filter(
        employee=user_profile,
        date__gte=prev_week_start,
        date__lt=week_start
    )
    prev_avail_list = []
    for a in prev_avail:
        # Сдвигаем дату на неделю вперёд
        new_date = a.date + timedelta(weeks=1)
        prev_avail_list.append({
            'date': new_date.strftime('%Y-%m-%d'),
            'time': a.start_time.strftime('%H:%M')
        })

    prev_week = (week_start - timedelta(weeks=1)).strftime('%Y-%m-%d')
    next_week = (week_start + timedelta(weeks=1)).strftime('%Y-%m-%d')

    context = {
        'days': current_days,
        'date_strings': date_strings,
        'slots': slots,
        'checked_keys': checked_keys,
        'last_updated': last_updated,
        'week_start': week_start,
        'week_end': week_start + timedelta(days=6),
        'prev_week': prev_week,
        'next_week': next_week,
        'prev_avail_json': json.dumps(prev_avail_list),
    }
    return render(request, 'core/availability/my_availability.html', context)


#для отправки напоминаний о доступности
@login_required
@user_passes_test(lambda u: u.profile.role == 'manager')
def send_availability_reminder_manual(request):
    if request.method == "POST":
        employees = UserProfile.objects.filter(role='employee')
        emails = [emp.user.email for emp in employees if emp.user.email]
        if emails:
            ok = send_mail_with_fallback(
                subject="Напоминание: укажите ваши рабочие часы",
                message="Пожалуйста, зайдите в личный кабинет и укажите, когда вы можете работать на следующей неделе.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=emails,
            )
            if ok:
                messages.success(request, f"Напоминание отправлено {len(emails)} сотрудникам.")
            else:
                messages.error(
                    request,
                    "Письма не отправлены: почтовый сервер недоступен или отклоняет подключение. "
                    "Проверьте настройки SMTP и пароль приложения."
                )
        else:
            messages.warning(request, "Нет сотрудников с email.")
    return redirect('schedule_view')



""" === ОТЧЕТЫ === """
import json
from datetime import date, datetime, timedelta
from collections import defaultdict
import statistics
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import ShiftAssignment, UserProfile, WorkoutType, HourRateChange

@login_required
def reports_view(request):
    user = request.user
    profile = user.profile

    # === Detect current user role ===
    is_manager = (profile.role == 'manager')

    # Если не менеджер — показываем только его данные
    if not is_manager:
        employee_filter = profile  # текущий пользователь
    else:
        employee_id = request.GET.get('employee')
        if employee_id and employee_id != 'all':
            try:
                employee_filter = UserProfile.objects.get(id=employee_id, role='employee')
            except UserProfile.DoesNotExist:
                employee_filter = None
        else:
            employee_filter = None  # все сотрудники

    latest_rate = HourRateChange.objects.order_by('-effective_from', '-id').first()
    hour_rate = float(latest_rate.rate) if latest_rate else None

    # Изменение ставки: только руководитель.
    if 'set_hour_rate' in request.GET:
        if not is_manager:
            messages.error(request, "Изменять часовую ставку может только руководитель.")
        else:
            new_rate_str = request.GET.get('hour_rate', '').strip()
            if new_rate_str:
                try:
                    new_rate = float(new_rate_str.replace(',', '.'))
                    if new_rate >= 0:
                        HourRateChange.objects.create(
                            rate=new_rate,
                            effective_from=timezone.now(),
                            changed_by=request.user,
                        )
                        hour_rate = new_rate
                        messages.success(
                            request,
                            f"Ставка изменена на {int(new_rate) if new_rate.is_integer() else new_rate} ₽/час. "
                            "Новая ставка применяется только к будущим сменам."
                        )
                    else:
                        messages.error(request, "Ставка не может быть отрицательной.")
                except ValueError:
                    messages.error(request, "Введите корректное число.")

    # === PERIOD ===
    period = request.GET.get('period', 'month')
    # === Resolve period ===
    today = date.today()

    # По умолчанию — текущий месяц
    default_start = today.replace(day=1)
    if today.month == 12:
        default_end = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
    else:
        default_end = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    # Получаем параметры из запроса
    period = request.GET.get('period', 'month')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Initialize dates with defaults
    start_date = default_start
    end_date = default_end

    # Обработка кастомного периода
    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            period = 'custom'
        except ValueError:
            messages.error(request, "Неверный формат даты. Используйте формат ГГГГ-ММ-ДД.")
    elif period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    # else: остаётся 'month' -> уже задан как default

    delta = end_date - start_date
    all_dates = [start_date + timedelta(days=i) for i in range(delta.days + 1)]

    # === Apply custom date range (if provided) ===
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if start_date_str and end_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            period = 'custom'  # чтобы отличать от week/month/year
        except ValueError:
            messages.error(request, "Неверный формат даты. Используйте формат ГГГГ-ММ-ДД.")
            start_date = None
            end_date = None
    else:
        # Старая логика для period
        today = date.today()
        if period == 'week':
            start_date = today - timedelta(days=7)
            end_date = today
        elif period == 'month':
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        elif period == 'year':
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
        else:
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)



    # === Нормализация периода (фикс залипания custom-дат) ===
    # В форме скрытые start_date/end_date отправляются всегда, даже когда выбран week/month/year.
    # Поэтому учитываем их только если period=custom.
    today = date.today()
    period = request.GET.get('period', 'month')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    elif period == 'custom':
        try:
            if not start_date_str or not end_date_str:
                raise ValueError("missing_custom_dates")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            messages.error(request, "Для произвольного периода укажите корректные даты 'С' и 'По'.")
            period = 'month'
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:
        # month (по умолчанию)
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    delta = end_date - start_date
    all_dates = [start_date + timedelta(days=i) for i in range(delta.days + 1)]

    # === Manager filters ===
    workout_id = request.GET.get('workout')
    search_query = request.GET.get('search', '').strip()
    coverage_direction = request.GET.get('coverage_direction', 'all')
    coverage_sort = request.GET.get('coverage_sort', 'trainers_desc')

    # === Filter application ===
    if is_manager:
        all_employees = UserProfile.objects.filter(role='employee').order_by('user__username')
        if employee_filter:
            employees = [employee_filter]
        else:
            employees = all_employees
    else:
        # Сотрудник — только он сам
        all_employees = [profile]
        employees = [profile]

    # === SHIFT QUERY ===
    assignments_base = ShiftAssignment.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('employee', 'workout_type')
    assignments = assignments_base

    # Применяем фильтры
    if is_manager:
        if employee_filter:
            assignments = assignments.filter(employee=employee_filter)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)
        if search_query:
            assignments = assignments.filter(
                Q(employee__user__username__icontains=search_query) |
                Q(workout_type__name__icontains=search_query)
            )
    else:
        # Сотрудник: только свои + фильтр по типу занятия
        assignments = assignments.filter(employee=profile)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)
        # Поиск не нужен — только свои данные

    # === Aggregation ===
    period_end_moment = datetime.combine(end_date, datetime.max.time())
    if timezone.is_naive(period_end_moment):
        period_end_moment = timezone.make_aware(period_end_moment, timezone.get_current_timezone())
    rate_changes = list(
        HourRateChange.objects.filter(effective_from__lte=period_end_moment)
        .order_by('-effective_from', '-id')
        .values_list('effective_from', 'rate')
    )
    rate_changes = [(dt, float(rate)) for dt, rate in rate_changes]

    from collections import defaultdict
    data = defaultdict(lambda: defaultdict(float))
    emp_hours = defaultdict(float)
    emp_salary = defaultdict(float)
    day_hours = [0.0] * 7
    workout_hours = defaultdict(float)
    date_hours = defaultdict(float)
    salary_available = False

    for a in assignments:
        if a.start_time is None or a.end_time is None:
            continue
        dur_raw = (datetime.combine(date.min, a.end_time) - datetime.combine(date.min, a.start_time)).total_seconds() / 3600
        dur = _round_half_up_to_int(dur_raw)
        data[a.employee_id][a.date] += dur
        emp_hours[a.employee.user.username] += dur
        day_hours[a.date.weekday()] += dur
        workout_hours[a.workout_type.name] += dur
        date_hours[a.date] += dur
        rate = _resolve_hour_rate_for_shift(a.date, a.start_time, rate_changes)
        if rate is not None:
            emp_salary[a.employee_id] += dur * rate
            salary_available = True

    total_hours = {}
    total_shifts = {}
    for emp in employees:
        emp_id = emp.id
        hours = sum(data[emp_id].values())
        shifts = len([h for h in data[emp_id].values() if h > 0])
        total_hours[emp.id] = round(hours, 2)
        total_shifts[emp.id] = shifts

    daily_totals = []
    for d in all_dates:
        total = sum(data[emp.id].get(d, 0) for emp in employees)
        daily_totals.append(int(total))

    total_salary_per_emp = {}
    for emp in employees:
        total_salary_per_emp[emp.id] = int(round(emp_salary.get(emp.id, 0)))

    total_salary = int(round(sum(total_salary_per_emp.values())))

    # === Charts ===
    chart_data = {
        'empNames': list(emp_hours.keys()) or [],
        'empValues': [round(v, 2) for v in emp_hours.values()] or [],
        'dayLabels': ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс'],
        'dayValues': [round(v, 2) for v in day_hours] or [0]*7,
        'workoutLabels': list(workout_hours.keys()) or [],
        'workoutValues': [round(v, 2) for v in workout_hours.values()] or [],
        'dateLabels': [d.strftime('%d.%m') for d in sorted(date_hours.keys())] or [],
        'dateValues': [round(date_hours[d], 2) for d in sorted(date_hours.keys())] or [],
    }

    # === Direction summary for manager ===
    direction_rows = []
    employees_direction_rows = []
    direction_summary = {}
    workout_types_all = WorkoutType.objects.all().order_by('name')
    staff_rows = []
    staff_sort = request.GET.get('staff_sort', 'hours_desc')
    staff_peak_filter = request.GET.get('staff_peak_day', 'all')
    staff_swap_filter = request.GET.get('staff_swap_level', 'all')
    staff_search_raw = request.GET.get('staff_search', '').strip()
    staff_search = staff_search_raw.lower()

    if is_manager:
        def _display_name(user_profile):
            user_obj = user_profile.user
            full_name = f"{user_obj.last_name} {user_obj.first_name} {user_profile.patronymic}".strip()
            return full_name if full_name else user_obj.username

        all_employee_profiles = UserProfile.objects.filter(
            role='employee'
        ).select_related('user', 'employee_profile').prefetch_related('employee_profile__workout_types')

        # Кто какие направления ведет
        direction_to_trainers = {wt.id: [] for wt in workout_types_all}
        employees_without_directions = []

        for emp_profile in all_employee_profiles:
            emp_obj = getattr(emp_profile, 'employee_profile', None)
            if not emp_obj:
                continue

            trainer_name = _display_name(emp_profile)
            trainer_workouts = list(emp_obj.workout_types.all())

            if not trainer_workouts:
                employees_without_directions.append(trainer_name)

            employees_direction_rows.append({
                'name': trainer_name,
                'workout_names': [w.name for w in trainer_workouts],
                'workout_count': len(trainer_workouts),
            })

            for wt in trainer_workouts:
                direction_to_trainers.setdefault(wt.id, []).append(trainer_name)

        # Нагрузка по направлениям за выбранный период
        direction_assignments = assignments_base
        if employee_filter:
            direction_assignments = direction_assignments.filter(employee=employee_filter)

        direction_hours = defaultdict(float)
        direction_shifts = defaultdict(int)
        for shift in direction_assignments:
            if not shift.workout_type_id:
                continue
            if shift.start_time is None or shift.end_time is None:
                continue
            duration_raw = (
                datetime.combine(date.min, shift.end_time) -
                datetime.combine(date.min, shift.start_time)
            ).total_seconds() / 3600
            duration = _round_half_up_to_int(duration_raw)
            direction_hours[shift.workout_type_id] += duration
            direction_shifts[shift.workout_type_id] += 1

        for wt in workout_types_all:
            if coverage_direction != 'all' and str(wt.id) != str(coverage_direction):
                continue
            trainers = sorted(direction_to_trainers.get(wt.id, []))
            direction_rows.append({
                'id': wt.id,
                'name': wt.name,
                'trainers': trainers,
                'trainers_count': len(trainers),
                'hours': round(direction_hours.get(wt.id, 0), 2),
                'shifts': direction_shifts.get(wt.id, 0),
            })

        if coverage_sort == 'trainers_asc':
            direction_rows.sort(key=lambda x: (x['trainers_count'], x['name'].lower()))
        elif coverage_sort == 'hours_desc':
            direction_rows.sort(key=lambda x: (-x['hours'], x['name'].lower()))
        elif coverage_sort == 'hours_asc':
            direction_rows.sort(key=lambda x: (x['hours'], x['name'].lower()))
        elif coverage_sort == 'name_asc':
            direction_rows.sort(key=lambda x: x['name'].lower())
        elif coverage_sort == 'name_desc':
            direction_rows.sort(key=lambda x: x['name'].lower(), reverse=True)
        else:
            direction_rows.sort(key=lambda x: (-x['trainers_count'], x['name'].lower()))

        covered_count = len([r for r in direction_rows if r['trainers_count'] > 0])
        direction_summary = {
            'total_directions': len(direction_rows),
            'covered_directions': covered_count,
            'uncovered_directions': len(direction_rows) - covered_count,
            'employees_without_directions_count': len(employees_without_directions),
            'employees_without_directions': employees_without_directions,
        }

        # === Staff analytics ===
        weekday_names = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']

        assignments_for_staff = assignments_base
        if employee_filter:
            assignments_for_staff = assignments_for_staff.filter(employee=employee_filter)
        if workout_id and workout_id != 'all':
            assignments_for_staff = assignments_for_staff.filter(workout_type_id=workout_id)

        daily_hours_map = defaultdict(lambda: defaultdict(float))
        shifts_count_map = defaultdict(int)
        for shift in assignments_for_staff:
            if shift.start_time is None or shift.end_time is None:
                continue
            duration_raw = (
                datetime.combine(date.min, shift.end_time) -
                datetime.combine(date.min, shift.start_time)
            ).total_seconds() / 3600
            duration = _round_half_up_to_int(duration_raw)
            daily_hours_map[shift.employee_id][shift.date] += duration
            shifts_count_map[shift.employee_id] += 1

        period_schedules = Schedule.objects.filter(
            start_date__lte=end_date,
            end_date__gte=start_date,
        )

        approvals_qs = ScheduleApproval.objects.filter(
            schedule__in=period_schedules
        ).exclude(approved__isnull=True).select_related('employee')

        approval_map = defaultdict(lambda: {'responded': 0, 'rejected': 0})
        for approval in approvals_qs:
            approval_map[approval.employee_id]['responded'] += 1
            if approval.approved is False:
                approval_map[approval.employee_id]['rejected'] += 1

        swap_qs = ShiftSwapRequest.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date,
            status__in=['approved_by_manager', 'completed']
        ).select_related('from_employee__user_profile', 'to_employee__user_profile')

        swap_involved_map = defaultdict(int)
        swap_sent_map = defaultdict(int)
        swap_received_map = defaultdict(int)
        for swap in swap_qs:
            from_profile_id = swap.from_employee.user_profile_id if swap.from_employee_id else None
            to_profile_id = swap.to_employee.user_profile_id if swap.to_employee_id else None
            if from_profile_id:
                swap_involved_map[from_profile_id] += 1
                swap_sent_map[from_profile_id] += 1
            if to_profile_id:
                swap_involved_map[to_profile_id] += 1
                swap_received_map[to_profile_id] += 1

        for emp_profile in all_employee_profiles:
            emp_id = emp_profile.id
            full_name = _display_name(emp_profile)
            login_name = emp_profile.user.username

            daily_items = daily_hours_map.get(emp_id, {})
            total_hours_emp = round(sum(daily_items.values()), 2)
            shifts_emp = shifts_count_map.get(emp_id, 0)
            worked_days = [h for h in daily_items.values() if h > 0]
            worked_days_count = len(worked_days)
            avg_per_workday = round(total_hours_emp / worked_days_count, 2) if worked_days_count else 0.0

            if worked_days_count >= 2:
                mean_val = statistics.mean(worked_days)
                std_val = statistics.pstdev(worked_days)
                cv = (std_val / mean_val) if mean_val > 0 else 1.0
                uniformity_score = max(0, round(100 - min(100, cv * 100), 1))
            elif worked_days_count == 1:
                uniformity_score = 100.0
            else:
                uniformity_score = 0.0

            peak_day_name = '—'
            if daily_items:
                peak_day_index, _ = max(
                    ((d.weekday(), hrs) for d, hrs in daily_items.items()),
                    key=lambda pair: pair[1]
                )
                peak_day_name = weekday_names[peak_day_index]

            approved_swaps = swap_involved_map.get(emp_id, 0)
            swaps_sent = swap_sent_map.get(emp_id, 0)
            swaps_received = swap_received_map.get(emp_id, 0)
            swap_frequency_pct = round((approved_swaps / shifts_emp) * 100, 1) if shifts_emp else 0.0

            responded = approval_map[emp_id]['responded']
            rejected = approval_map[emp_id]['rejected']
            rejection_pct = round((rejected / responded) * 100, 1) if responded else 0.0

            staff_rows.append({
                'employee_id': emp_id,
                'name': full_name,
                'username': login_name,
                'hours': total_hours_emp,
                'worked_days': worked_days_count,
                'shifts': shifts_emp,
                'avg_per_day': avg_per_workday,
                'uniformity': uniformity_score,
                'rejection_pct': rejection_pct,
                'rejected_count': rejected,
                'responded_count': responded,
                'swap_count': approved_swaps,
                'swap_sent': swaps_sent,
                'swap_received': swaps_received,
                'swap_frequency': swap_frequency_pct,
                'peak_day': peak_day_name,
            })

        if staff_peak_filter != 'all':
            staff_rows = [row for row in staff_rows if row['peak_day'] == staff_peak_filter]

        if staff_swap_filter == 'none':
            staff_rows = [row for row in staff_rows if row['swap_count'] == 0]
        elif staff_swap_filter == 'low':
            staff_rows = [row for row in staff_rows if 0 < row['swap_frequency'] <= 20]
        elif staff_swap_filter == 'high':
            staff_rows = [row for row in staff_rows if row['swap_frequency'] > 20]

        if staff_search:
            staff_rows = [
                row for row in staff_rows
                if staff_search in row['name'].lower() or staff_search in row['username'].lower()
            ]

        if staff_sort == 'hours_asc':
            staff_rows.sort(key=lambda x: (x['hours'], x['name'].lower()))
        elif staff_sort == 'uniformity_desc':
            staff_rows.sort(key=lambda x: (-x['uniformity'], x['name'].lower()))
        elif staff_sort == 'uniformity_asc':
            staff_rows.sort(key=lambda x: (x['uniformity'], x['name'].lower()))
        elif staff_sort == 'reject_desc':
            staff_rows.sort(key=lambda x: (-x['rejection_pct'], x['name'].lower()))
        elif staff_sort == 'swap_desc':
            staff_rows.sort(key=lambda x: (-x['swap_count'], x['name'].lower()))
        elif staff_sort == 'name_asc':
            staff_rows.sort(key=lambda x: x['name'].lower())
        else:
            staff_rows.sort(key=lambda x: (-x['hours'], x['name'].lower()))

    staff_chart = {
        'labels': [row['name'] for row in staff_rows[:12]],
        'hours': [row['hours'] for row in staff_rows[:12]],
        'uniformity': [row['uniformity'] for row in staff_rows[:12]],
        'rejection': [row['rejection_pct'] for row in staff_rows[:12]],
        'swaps': [row['swap_count'] for row in staff_rows[:12]],
    }

    staff_summary = {
        'trainers_count': len(staff_rows),
        'total_hours': round(sum((row.get('hours') or 0) for row in staff_rows), 1),
        'avg_uniformity': round(
            (sum((row.get('uniformity') or 0) for row in staff_rows) / len(staff_rows)),
            1
        ) if staff_rows else 0.0,
        'avg_rejection_pct': round(
            (sum((row.get('rejection_pct') or 0) for row in staff_rows) / len(staff_rows)),
            1
        ) if staff_rows else 0.0,
        'total_swaps': int(sum((row.get('swap_count') or 0) for row in staff_rows)),
    }

    context = {
        'start_date': start_date,
    'end_date': end_date,
    'period': period,
        'employee_id': getattr(employee_filter, 'id', 'all') if is_manager else 'self',
        'workout_id': workout_id or 'all',
        'search_query': search_query,
        'employees': employees,
        'all_employees': all_employees if is_manager else [],
        'workout_types': WorkoutType.objects.all(),
        'all_dates': all_dates,
        'data': data,
        'total_hours': total_hours,
        'total_shifts': total_shifts,
        'daily_totals': daily_totals,
        'total_all_hours': round(sum(total_hours.values()), 2),
        'total_all_assignments': assignments.count(),
        'active_employees': len(employees),
        'chart_data_json': json.dumps(chart_data, ensure_ascii=False),
        'hour_rate': hour_rate,
        'rate_last_changed_at': latest_rate.effective_from if latest_rate else None,
        'total_salary': total_salary,
        'total_salary_per_emp': total_salary_per_emp,
        'salary_available': salary_available,
        'is_manager': is_manager,
        'coverage_direction': coverage_direction,
        'coverage_sort': coverage_sort,
        'direction_rows': direction_rows,
        'employees_direction_rows': employees_direction_rows,
        'direction_summary': direction_summary,
        'staff_rows': staff_rows,
        'staff_sort': staff_sort,
        'staff_peak_filter': staff_peak_filter,
        'staff_swap_filter': staff_swap_filter,
        'staff_search': staff_search_raw,
        'staff_chart_json': json.dumps(staff_chart, ensure_ascii=False),
        'staff_summary': staff_summary,
    }
    return render(request, 'core/reports/reports.html', context)



""" === EXPORT TO EXCEL === """
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import date, timedelta, datetime
from decimal import Decimal, ROUND_HALF_UP
from .models import ShiftAssignment, UserProfile, HourRateChange

def _format_number(value):
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return round(value, 1) if isinstance(value, float) else value


def _round_half_up_to_int(value):
    """Округление до целого по правилам half-up (2.5 -> 3)."""
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def _resolve_hour_rate_for_shift(shift_date, shift_start_time, rate_changes):
    """
    Возвращает часовую ставку, которая действовала на момент начала смены.
    rate_changes: список кортежей (effective_from, rate), отсортированный по убыванию effective_from.
    """
    if not rate_changes:
        return None

    moment = datetime.combine(shift_date, shift_start_time or datetime.min.time())
    if timezone.is_naive(moment):
        moment = timezone.make_aware(moment, timezone.get_current_timezone())

    for effective_from, rate in rate_changes:
        if effective_from <= moment:
            return float(rate)
    return None

@login_required
def export_operational_excel(request):
    today = date.today()  # <- initialized

    period = request.GET.get('period', 'month')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    # Логика периода = как на странице отчетов.
    if period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    elif period == 'custom':
        try:
            if not start_date_str or not end_date_str:
                raise ValueError("missing_custom_dates")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:  # month
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)


    # === Build full date range ===
    all_dates = []
    d = start_date
    while d <= end_date:
        all_dates.append(d)
        d += timedelta(days=1)

    is_manager = request.user.profile.role == 'manager'
    employee_id = request.GET.get('employee')
    workout_id = request.GET.get('workout')
    search_query = request.GET.get('search', '').strip()

    # Список сотрудников (как в отчете на странице)
    if is_manager:
        all_employees_qs = UserProfile.objects.filter(role='employee').order_by('user__username')
        employee_filter = None
        if employee_id and employee_id != 'all':
            employee_filter = all_employees_qs.filter(id=employee_id).first()
        employees = [employee_filter] if employee_filter else list(all_employees_qs)
    else:
        employee_filter = request.user.profile
        employees = [request.user.profile]

    assignments = ShiftAssignment.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('employee')

    # Фильтрация назначений = как на странице отчетов.
    if is_manager:
        if employee_filter:
            assignments = assignments.filter(employee=employee_filter)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)
        if search_query:
            assignments = assignments.filter(
                Q(employee__user__username__icontains=search_query) |
                Q(workout_type__name__icontains=search_query)
            )
    else:
        assignments = assignments.filter(employee=request.user.profile)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)

    from collections import defaultdict
    data = defaultdict(lambda: defaultdict(float))
    for a in assignments:
        if a.start_time is None or a.end_time is None:
            continue  # пропускаем смены без времени
        dur_raw = (datetime.combine(date.min, a.end_time) - datetime.combine(date.min, a.start_time)).total_seconds() / 3600
        dur = _round_half_up_to_int(dur_raw)
        data[a.employee_id][a.date] += dur

    period_end_moment = datetime.combine(end_date, datetime.max.time())
    if timezone.is_naive(period_end_moment):
        period_end_moment = timezone.make_aware(period_end_moment, timezone.get_current_timezone())
    rate_changes = list(
        HourRateChange.objects.filter(effective_from__lte=period_end_moment)
        .order_by('-effective_from', '-id')
        .values_list('effective_from', 'rate')
    )
    rate_changes = [(dt, float(rate)) for dt, rate in rate_changes]

    # Подсчёт
    total_hours_per_emp = {}
    total_salary_per_emp = {}
    salary_by_emp = defaultdict(float)
    for a in assignments:
        if a.start_time is None or a.end_time is None:
            continue
        dur_raw = (datetime.combine(date.min, a.end_time) - datetime.combine(date.min, a.start_time)).total_seconds() / 3600
        dur = _round_half_up_to_int(dur_raw)
        rate = _resolve_hour_rate_for_shift(a.date, a.start_time, rate_changes)
        if rate is not None:
            salary_by_emp[a.employee_id] += dur * rate

    for emp in employees:
        emp_id = emp.id
        hours = sum(data[emp_id].values())
        total_hours_per_emp[emp.id] = _format_number(hours)
        salary = salary_by_emp.get(emp.id, 0)
        total_salary_per_emp[emp.id] = _format_number(salary)

    # === EXCEL ===
    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    # Граница: все ячейки — тонкая рамка
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Ширина столбцов
    ws.column_dimensions['A'].width = 25          # Сотрудник — широкий
    ws.column_dimensions['B'].width = 26          # Период
    ws.column_dimensions['C'].width = 10          # ЗП
    for i in range(len(all_dates)):
        col_letter = get_column_letter(4 + i)      # D = 4
        ws.column_dimensions[col_letter].width = 6  # Узкие дни, но чуть шире для читаемости

    # Заголовки (строка 1)
    ws.cell(row=1, column=1, value="Сотрудник")
    period_label = f"{start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}"
    ws.cell(row=1, column=2, value=period_label)
    ws.cell(row=1, column=3, value="ЗП")

    # Даты: "01.01", "02.01"...
    for i, d in enumerate(all_dates, start=4):  # D = 4
        ws.cell(row=1, column=i, value=f"{d.day:02d}.{d.month:02d}")

    # Жирный шрифт для всей первой строки
    bold_font = Font(bold=True)
    for col in range(1, 4 + len(all_dates)):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.border = thin_border

    # Заливка заголовков B и C — зелёная
    green_fill = PatternFill(start_color="0099FF00", end_color="0099FF00", fill_type="solid")
    for col in [2, 3]:
        ws.cell(row=1, column=col).fill = green_fill

    def _display_name(profile: UserProfile) -> str:
        user_obj = profile.user
        full_name = f"{user_obj.last_name} {user_obj.first_name} {profile.patronymic or ''}".strip()
        return full_name if full_name else user_obj.username

    # Данные по сотрудникам
    for row_idx, emp in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=_display_name(emp))
        ws.cell(row=row_idx, column=2, value=total_hours_per_emp[emp.id])
        ws.cell(row=row_idx, column=3, value=total_salary_per_emp[emp.id])

        # Зелёная заливка для B и C
        for col in [2, 3]:
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = green_fill
            cell.border = thin_border

        # Дни: без нулей, с границами
        for i, d in enumerate(all_dates, start=4):
            h = data[emp.id].get(d, 0)
            val = _format_number(h) if h != 0 else ""
            cell = ws.cell(row=row_idx, column=i, value=val)
            cell.border = thin_border

    # Summary row
    last_row = len(employees) + 2
    ws.cell(row=last_row, column=1, value="Итого кол-часов")

    total_hours_all = sum(total_hours_per_emp.values())
    total_salary_all = sum(total_salary_per_emp.values())

    ws.cell(row=last_row, column=2, value=_format_number(total_hours_all))
    ws.cell(row=last_row, column=3, value=_format_number(total_salary_all))

    # Дни — сумма
    for i, d in enumerate(all_dates, start=4):
        total_day = sum(data[emp.id].get(d, 0) for emp in employees)
        val = _format_number(total_day) if total_day != 0 else ""
        cell = ws.cell(row=last_row, column=i, value=val)
        cell.border = thin_border

    # Форматирование итоговой строки
    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    purple_font = Font(color="800080", bold=True)

    for col in range(1, 4 + len(all_dates)):
        cell = ws.cell(row=last_row, column=col)
        cell.fill = yellow_fill
        cell.font = purple_font
        cell.border = thin_border

    # Применяем границы ко ВСЕМ ячейкам таблицы (включая внутренние)
    max_row = last_row
    max_col = 3 + len(all_dates)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.border:
                cell.border = thin_border

    ws.freeze_panes = "D2"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"tabel_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
