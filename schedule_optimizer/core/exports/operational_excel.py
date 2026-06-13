"""Формирует Excel-табель по сменам, часам и зарплате за выбранный период."""

from collections import defaultdict
from datetime import date, datetime, timedelta
from decimal import Decimal, ROUND_HALF_UP

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..models import HourRateChange, ShiftAssignment, UserProfile


def _format_number(value):
    """Приводит число к удобному виду для вывода в отчете или Excel-файле."""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return round(value, 1) if isinstance(value, float) else value


def _round_half_up_to_int(value):
    """Округляет часы до целого значения по правилу half-up, чтобы 2.5 стало 3."""
    return int(Decimal(str(value)).quantize(Decimal('1'), rounding=ROUND_HALF_UP))


def _resolve_hour_rate_for_shift(shift_date, shift_start_time, rate_changes):
    """Находит часовую ставку, которая действовала на момент начала конкретной смены."""
    if not rate_changes:
        return None

    moment = datetime.combine(shift_date, shift_start_time or datetime.min.time())
    if timezone.is_naive(moment):
        moment = timezone.make_aware(moment, timezone.get_current_timezone())

    for effective_from, rate in rate_changes:
        if effective_from <= moment:
            return float(rate)
    return None


def _display_name(profile: UserProfile) -> str:
    """Собирает читаемое имя пользователя из фамилии, имени, отчества или логина."""
    user_obj = profile.user
    full_name = f"{user_obj.last_name} {user_obj.first_name} {profile.patronymic or ''}".strip()
    return full_name if full_name else user_obj.username


@login_required
def export_operational_excel(request):
    """Формирует и отдает Excel-файл табеля за выбранный период."""
    today = date.today()
    period = request.GET.get('period', 'month')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')

    if period == 'week':
        start_date = today - timedelta(days=7)
        end_date = today
    elif period == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    elif period == 'custom':
        try:
            if not start_date_str or not end_date_str:
                raise ValueError("missing_custom_dates")
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if start_date > end_date:
                start_date, end_date = end_date, start_date
        except ValueError:
            start_date = today.replace(day=1)
            if today.month == 12:
                end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
            else:
                end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
    else:
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_date = today.replace(month=today.month + 1, day=1) - timedelta(days=1)

    all_dates = []
    d = start_date
    while d <= end_date:
        all_dates.append(d)
        d += timedelta(days=1)

    is_manager = request.user.profile.role == 'manager'
    employee_id = request.GET.get('employee')
    workout_id = request.GET.get('workout')
    search_query = request.GET.get('search', '').strip()

    if is_manager:
        all_employees_qs = UserProfile.objects.filter(role='employee').order_by('user__username')
        employee_filter = None
        if employee_id and employee_id != 'all':
            employee_filter = all_employees_qs.filter(id=employee_id).first()
        employees = [employee_filter] if employee_filter else list(all_employees_qs)
    else:
        employee_filter = request.user.profile
        employees = [request.user.profile]

    assignments = ShiftAssignment.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
    ).select_related('employee')

    if is_manager:
        if employee_filter:
            assignments = assignments.filter(employee=employee_filter)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)
        if search_query:
            assignments = assignments.filter(
                Q(employee__user__username__icontains=search_query)
                | Q(workout_type__name__icontains=search_query)
            )
    else:
        assignments = assignments.filter(employee=request.user.profile)
        if workout_id and workout_id != 'all':
            assignments = assignments.filter(workout_type_id=workout_id)

    data = defaultdict(lambda: defaultdict(float))
    for assignment in assignments:
        if assignment.start_time is None or assignment.end_time is None:
            continue
        dur_raw = (
            datetime.combine(date.min, assignment.end_time)
            - datetime.combine(date.min, assignment.start_time)
        ).total_seconds() / 3600
        dur = _round_half_up_to_int(dur_raw)
        data[assignment.employee_id][assignment.date] += dur

    period_end_moment = datetime.combine(end_date, datetime.max.time())
    if timezone.is_naive(period_end_moment):
        period_end_moment = timezone.make_aware(period_end_moment, timezone.get_current_timezone())

    rate_changes = list(
        HourRateChange.objects.filter(effective_from__lte=period_end_moment)
        .order_by('-effective_from', '-id')
        .values_list('effective_from', 'rate')
    )
    rate_changes = [(dt, float(rate)) for dt, rate in rate_changes]

    total_hours_per_emp = {}
    total_salary_per_emp = {}
    salary_by_emp = defaultdict(float)
    for assignment in assignments:
        if assignment.start_time is None or assignment.end_time is None:
            continue
        dur_raw = (
            datetime.combine(date.min, assignment.end_time)
            - datetime.combine(date.min, assignment.start_time)
        ).total_seconds() / 3600
        dur = _round_half_up_to_int(dur_raw)
        rate = _resolve_hour_rate_for_shift(assignment.date, assignment.start_time, rate_changes)
        if rate is not None:
            salary_by_emp[assignment.employee_id] += dur * rate

    for employee in employees:
        hours = sum(data[employee.id].values())
        total_hours_per_emp[employee.id] = _format_number(hours)
        total_salary_per_emp[employee.id] = _format_number(salary_by_emp.get(employee.id, 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Табель"

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 10
    for i in range(len(all_dates)):
        col_letter = get_column_letter(4 + i)
        ws.column_dimensions[col_letter].width = 6

    ws.cell(row=1, column=1, value="Сотрудник")
    period_label = f"{start_date.strftime('%d.%m.%Y')} – {end_date.strftime('%d.%m.%Y')}"
    ws.cell(row=1, column=2, value=period_label)
    ws.cell(row=1, column=3, value="ЗП")

    for i, day in enumerate(all_dates, start=4):
        ws.cell(row=1, column=i, value=f"{day.day:02d}.{day.month:02d}")

    bold_font = Font(bold=True)
    for col in range(1, 4 + len(all_dates)):
        cell = ws.cell(row=1, column=col)
        cell.font = bold_font
        cell.border = thin_border

    green_fill = PatternFill(start_color="0099FF00", end_color="0099FF00", fill_type="solid")
    for col in [2, 3]:
        ws.cell(row=1, column=col).fill = green_fill

    for row_idx, employee in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=_display_name(employee))
        ws.cell(row=row_idx, column=2, value=total_hours_per_emp[employee.id])
        ws.cell(row=row_idx, column=3, value=total_salary_per_emp[employee.id])

        for col in [2, 3]:
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = green_fill
            cell.border = thin_border

        for i, day in enumerate(all_dates, start=4):
            hours = data[employee.id].get(day, 0)
            value = _format_number(hours) if hours != 0 else ""
            cell = ws.cell(row=row_idx, column=i, value=value)
            cell.border = thin_border

    last_row = len(employees) + 2
    ws.cell(row=last_row, column=1, value="Итого кол-часов")
    ws.cell(row=last_row, column=2, value=_format_number(sum(total_hours_per_emp.values())))
    ws.cell(row=last_row, column=3, value=_format_number(sum(total_salary_per_emp.values())))

    for i, day in enumerate(all_dates, start=4):
        total_day = sum(data[employee.id].get(day, 0) for employee in employees)
        value = _format_number(total_day) if total_day != 0 else ""
        cell = ws.cell(row=last_row, column=i, value=value)
        cell.border = thin_border

    yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    purple_font = Font(color="800080", bold=True)

    for col in range(1, 4 + len(all_dates)):
        cell = ws.cell(row=last_row, column=col)
        cell.fill = yellow_fill
        cell.font = purple_font
        cell.border = thin_border

    max_row = last_row
    max_col = 3 + len(all_dates)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.border:
                cell.border = thin_border

    ws.freeze_panes = "D2"

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"tabel_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
